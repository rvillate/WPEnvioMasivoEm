import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


from log import log_print

# Diccionario global para almacenar los datos en memoria por sesión
excel_sessions = {}

def write_log_to_excel(file_path, modulo, seccion, accion, estado, comentario):
    """
    Escribe un registro en un archivo Excel en la última fila disponible.
    Si hay un error al escribir, guarda el error en un log secundario en formato .log.
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        last_row = ws.max_row + 1
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.append([fecha_hora, modulo, seccion, accion, estado, comentario])

        wb.save(file_path)
        return True, "Log guardado exitosamente."

    except Exception as e:
        error_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Error al escribir en Excel: {e}\n"
        save_error_to_log(error_message)
        print(f"Error al escribir en el log de Excel: {e}")
        return False, f"Error al escribir en el log de Excel: {e}"


def load_excel_data_to_memory(file_path, sheet_name, session_name):
    """
    Carga los datos de una hoja específica de un archivo Excel en memoria y los asocia a una sesión con nombre.

    :param file_path: Ruta del archivo Excel
    :param sheet_name: Nombre de la hoja del Excel a cargar
    :param session_name: Nombre identificador de la sesión
    :return: (bool, str) Resultado y mensaje
    """
    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe en el archivo."

        ws = wb[sheet_name]  # Usar la hoja especificada

        headers = [cell.value for cell in ws[1]]
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        excel_sessions[session_name] = data
        return True, f"Datos de la hoja '{sheet_name}' cargados en sesión '{session_name}' exitosamente."

    except Exception as e:
        error_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Error al leer Excel en sesión '{session_name}': {e}\n"
        save_error_to_log(error_message)
        print(f"Error al leer Excel en sesión '{session_name}': {e}")
        return False, f"Error al leer Excel en sesión '{session_name}': {e}"


def update_cell_excel_by_column_name(file_path, sheet_name, column_name, row, value):
    """
    Actualiza una celda en un archivo Excel, identificando la columna por nombre.

    :param file_path: Ruta al archivo Excel
    :param sheet_name: Nombre de la hoja
    :param column_name: Nombre de la columna a modificar
    :param row: Número de fila (empieza desde 2, porque 1 es el encabezado)
    :param value: Valor que se quiere escribir
    :return: (bool, str) Resultado y mensaje
    """
    try:
        row = row + 1

        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe en el archivo."

        ws = wb[sheet_name]

        # Buscar la columna por nombre en la fila 1
        col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == column_name:
                col_index = idx
                break

        if col_index is None:
            return False, f"No se encontró la columna con nombre '{column_name}'."

        # Actualizar la celda
        ws.cell(row=row, column=col_index, value=value)

        # Guardar el archivo
        wb.save(file_path)
        return True, f"Celda en fila {row}, columna '{column_name}' actualizada con valor '{value}'."

    except Exception as e:
        error_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Error al actualizar Excel: {e}\n"
        save_error_to_log(error_message)
        return False, f"Error al actualizar Excel: {e}"


def get_excel_cell(session_name, row_number, column_name):
    """
    Obtiene el valor de una celda desde una sesión Excel cargada previamente en memoria.

    :param session_name: Nombre de la sesión previamente cargada
    :param row_number: Número de fila (comenzando desde 1 para la fila de encabezados)
    :param column_name: Nombre de la columna (corresponde al encabezado)
    :return: Valor de la celda o cadena vacía si no existe
    """
    session_data = excel_sessions.get(session_name)
    if not session_data:
        log_print("No hay sesiones almacenadas para la recuperación de datos")
        return ("No hay sesiones almacenadas para la recuperación de datos", False, "")

    if row_number-1 < 0 or row_number-1 >= len(session_data):
        return ("No hay data en la sesion almacenada", False, "")

    row_data = session_data[row_number-1]
    return "Data recuperada correctamente", True, (row_data.get(column_name, ""))


def save_error_to_log(error_message):
    """
    Guarda errores en un archivo .log con formato 'log_dd-MM-yy.log'.

    :param error_message: Mensaje de error a guardar en el log
    """
    try:
        log_filename = f"log_{datetime.now().strftime('%d-%m-%y')}.log"
        log_path = os.path.join(os.getcwd(), log_filename)

        with open(log_path, "a", encoding="utf-8") as log_file:
            log_file.write(error_message)
        print(log_path)
    except Exception as log_error:
        print(f"Error crítico al escribir en el archivo log: {log_error}")